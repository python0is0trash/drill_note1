from openpyxl import Workbook
from datetime import date

from openpyxl.styles import (
    Alignment, Font, NamedStyle
)

workbook = Workbook()
worksheet = workbook.active

workbook_alignment_center = Alignment(
    horizontal='center',
    vertical='center'
)
workbook_alignment_left = Alignment(
    horizontal='left',
    vertical='center'
)

main_style = NamedStyle(name="main_style")
main_style.font = Font(
    name='Times New Roman',
    size=48,
    bold=False,
    italic=False,
    underline='none',
    strike=False,
    color='00000000'
)
workbook.add_named_style(main_style)

## создание листов
worksheet_1 = workbook.create_sheet("List 1", 0)
worksheet_2 = workbook.create_sheet("List 2", 1)
if 'Sheet' in workbook.sheetnames:
    workbook.remove(workbook["Sheet"])

month_list = [
    'января',
    'февраля',
    'марта',
    'апреля',
    'мая',
    'июня',
    'июля',
    'августа',
    'сентября',
    'октября',
    'ноября',
    'декабря'
]

date_today = [date.today().day, month_list[date.today().month - 1], date.today().year]
help_list = [
    'Причины отсутствия',
    'курсантов (слушателей) ФПСОИБ',
    f'{date_today[0]} {date_today[1]} {date_today[2]} г.',
    'УУМР (факс): 8(495)779-23-92;     Комендантский отдел т.(факс): 8(499)793-54-36 ',
    'Взвод',
    'Всего на взвод',
    'Всего',
    'БОЛЬНЫЕ',
    'НАРЯД',
    'ДРУГИЕ ПРИЧИНЫ',
    'Кол-во',
    'Причина',
    'Итого за факультет',
    'Начальник ФПСОИБ',
    'подполковник полиции',
    'Д.Л. Курякин'
]

# examples

info_list = [
    'Перваков И.И.(ОРВИ, 30.01.24), Чапля Н.Я.(Covid-19, 30.01.24), Грудачев К.С.(ОРВИ, 30.01.24), Хмелев В.С.(ОРВИ, 01.02.24)',
    'Репещук И.В.(Н)',
    'Мягков С.В., Умяров К.В.'
    # 'Направление к врачу, 02.02.24'
]

total_count = {
    'ill': {

    },
    'working': {},
    'other': {}
}

platoons = [
    '931',
    '932',
    '933',
    '922',
    '923',
    '912'
]
